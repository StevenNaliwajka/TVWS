import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def build_tof_template_xlsx(
    out_path: Path,
    num_runs: int,
    edges_rows_per_run: int = 8,     # must be even (pairs)
    tof_centers_per_run: int = 4,    # 4 centers = 8 edge rows
    run_number_width: int = 4,
):
    if edges_rows_per_run % 2 != 0:
        raise ValueError("edges_rows_per_run must be even (pairs of edges).")

    wb = Workbook()
    ws = wb.active
    ws.title = "ToF"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # column widths similar to your layout
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 4
    for i in range(tof_centers_per_run):
        ws.column_dimensions[get_column_letter(5 + i)].width = 12  # E.. columns

    row = 1
    for r in range(1, num_runs + 1):
        run_label = f"run_{r:0{run_number_width}d}"

        # header row
        ws.cell(row=row, column=1, value=run_label).font = bold
        ws.cell(row=row, column=2, value="Edges").font = bold
        ws.cell(row=row, column=4, value="ToF Center").font = bold
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=4).alignment = center

        edge_start = row + 1

        # edge rows (you fill col B = edge time, col C = edge value)
        for i in range(edges_rows_per_run):
            ws.cell(row=edge_start + i, column=2, value=None)
            ws.cell(row=edge_start + i, column=3, value=None)

        # ToF center formulas written on the first edge row (edge_start), in E.. columns
        for j in range(tof_centers_per_run):
            r1 = edge_start + 2*j
            r2 = edge_start + 2*j + 1
            col = 5 + j  # E=5
            ws.cell(
                row=edge_start,
                column=col,
                value=f'=IF(OR(B{r1}="",B{r2}=""),"",AVERAGE(B{r1},B{r2}))'
            ).number_format = "0.00000"

        # blank spacer row
        row = edge_start + edges_rows_per_run + 2

    wb.save(out_path)
    print(f"[OK] wrote: {out_path}")

if __name__ == "__main__":
    out_name = input("Output XLSX filename (default: tof_template.xlsx): ").strip() or "tof_template.xlsx"
    n = int(input("How many runs will you be doing? ").strip())

    edges_rows = input("How many edge rows per run? (default 8): ").strip()
    edges_rows = int(edges_rows) if edges_rows else 8

    centers = input("How many ToF centers per run? (default 4): ").strip()
    centers = int(centers) if centers else 4

    build_tof_template_xlsx(Path(out_name), n, edges_rows_per_run=edges_rows, tof_centers_per_run=centers)