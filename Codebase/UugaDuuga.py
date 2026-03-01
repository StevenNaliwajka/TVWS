from importlib.metadata import metadata
import plotly.graph_objects as go
import plotly.io as pio
import numpy as np
import matplotlib.pyplot as plt
import json

from pathlib import Path
from scipy.signal import butter, filtfilt, spectrogram, find_peaks, firwin, lfilter
from Codebase.FileIO.collect_all_data import load_signal_grid
from scipy.ndimage import gaussian_filter1d, gaussian_laplace

from Codebase.Filter.filter_singal import filter_signal

from Codebase.Object.metadata_object import MetaDataObj
from Codebase.TOF.Type3.compute_relative_tof import compute_relative_tof
from Codebase.TOF.Type4.compute_tof import compute_tof
from Codebase.process_signal import process_signal

import tkinter as tk
from tkinter import filedialog
from pathlib import Path

import argparse
import csv

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from collections import defaultdict

'''
Current Wants:
    add file explorer functionality to remove need for copying file name
    research possible filtering techniques to improve edge detection
    make it iterable over entire data folder
    add excel output file that tracks ToF to specified folder of certain transmission distances
    automate the magnitude and cluster distance variables
    Get some sleep
'''

def save_interactive_html_clickpick(x, y, iq_path):
    """
    Interactive HTML:
      - Hover on signal works
      - Click adds multiple picked points (persist during session)
      - Picked points show hover labels
      - Shift+Click removes nearest picked point
      - Download picks CSV button
    """
    x = [float(v) for v in x]
    y = [float(v) for v in y]

    fig = go.Figure()

    # Signal trace (hover shows x,y)
    fig.add_trace(go.Scattergl(
        x=x, y=y,
        mode="lines",
        name="signal",
        hovertemplate="x=%{x}<br>y=%{y}<extra></extra>",
    ))

    # Picked points trace (empty at first)
    fig.add_trace(go.Scattergl(
        x=[], y=[],
        mode="markers",
        name="picked",
        marker=dict(size=10, symbol="x"),
        hovertemplate="PICK<br>x=%{x}<br>y=%{y}<extra></extra>",
    ))

    fig.update_layout(
        title=f"{iq_path.parent.name} / {iq_path.name}",
        xaxis_title="Time",
        yaxis_title="Magnitude",
        hovermode="closest",
    )

    storage_key = f"plotly_picks::{iq_path.stem}"

    post_js = f"""
    (function() {{
      const gd = document.querySelector('.plotly-graph-div');
      if (!gd) return;

      const storageKey = {json.dumps(storage_key)};
      let picks = [];

      function loadPicks() {{
        try {{
          const raw = localStorage.getItem(storageKey);
          if (!raw) return [];
          const arr = JSON.parse(raw);
          return Array.isArray(arr) ? arr : [];
        }} catch(e) {{
          return [];
        }}
      }}

      function savePicks(arr) {{
        try {{
          localStorage.setItem(storageKey, JSON.stringify(arr));
        }} catch(e) {{}}
      }}

      function redrawPicks(arr) {{
        const xs = arr.map(p => p.x);
        const ys = arr.map(p => p.y);
        Plotly.restyle(gd, {{ x: [xs], y: [ys] }}, [1]); // picked trace index = 1
      }}

      function appendPick(x, y) {{
        // Append visually (does NOT overwrite existing points)
        Plotly.extendTraces(gd, {{ x: [[x]], y: [[y]] }}, [1]);
      }}

      function nearestPickIndex(arr, x, y) {{
        if (!arr.length) return -1;
        let bestI = 0, bestD = Infinity;
        for (let i=0; i<arr.length; i++) {{
          const dx = arr[i].x - x;
          const dy = arr[i].y - y;
          const d = dx*dx + dy*dy;
          if (d < bestD) {{ bestD = d; bestI = i; }}
        }}
        return bestI;
      }}

      function downloadCSV(arr) {{
         // sort by x ascending (numeric)
          const sorted = arr.slice().sort((a, b) => a.x - b.x);
        
          const header = "x,y\\n";
          const rows = sorted.map(p => `${{p.x}},${{p.y}}`).join("\\n");
        
          const blob = new Blob([header + rows + "\\n"], {{type:"text/csv"}});
          const url = URL.createObjectURL(blob);
          const a = document.createElement("a");
          a.href = url;
          a.download = "YOURFILE_picks.csv";
          document.body.appendChild(a);
          a.click();
          a.remove();
          URL.revokeObjectURL(url);
      }}

      function addControlsOnce() {{
        if (gd.__hasPickControls) return;
        gd.__hasPickControls = true;

        const container = gd.parentElement;
        const bar = document.createElement("div");
        bar.style.display = "flex";
        bar.style.gap = "8px";
        bar.style.margin = "8px 0";
        bar.style.flexWrap = "wrap";

        const clearBtn = document.createElement("button");
        clearBtn.textContent = "Clear picks";
        clearBtn.onclick = () => {{
          picks = [];
          savePicks(picks);
          redrawPicks(picks);
        }};

        const dlBtn = document.createElement("button");
        dlBtn.textContent = "Download picks (CSV)";
        dlBtn.onclick = () => downloadCSV(picks);

        const note = document.createElement("div");
        note.textContent = "Click adds pick. Shift+Click removes nearest pick.";
        note.style.alignSelf = "center";
        note.style.opacity = "0.8";

        bar.appendChild(clearBtn);
        bar.appendChild(dlBtn);
        bar.appendChild(note);
        container.insertBefore(bar, gd);
      }}

      // Initialize from localStorage
      picks = loadPicks();
      addControlsOnce();
      redrawPicks(picks);

      // CLICK HANDLER:
      // IMPORTANT: Plotly click fires on whichever trace you clicked.
      // curveNumber 0 = signal, 1 = picked.
      gd.on('plotly_click', (ev) => {{
        if (!ev || !ev.points || ev.points.length === 0) return;

        const pt = ev.points[0];
        const x = pt.x;
        const y = pt.y;

        // Shift+click removes nearest picked point
        if (ev.event && ev.event.shiftKey) {{
          const i = nearestPickIndex(picks, x, y);
          if (i >= 0) {{
            picks.splice(i, 1);
            savePicks(picks);
            redrawPicks(picks);
          }}
          return;
        }}

        // Normal click: add pick
        picks.push({{x:x, y:y}});
        savePicks(picks);
        appendPick(x, y);
      }});
    }})();"""

    out_html = iq_path.parent / f"{iq_path.stem}_interactive.html"

    html = pio.to_html(
        fig,
        full_html=True,
        include_plotlyjs="cdn",
        config={"displaylogo": False},
        post_script=post_js
    )

    out_html.write_text(html, encoding="utf-8")
    return out_html
def build_reference_centers(centers_rows, capture_num: int):
    """
    Manually hard-set reference pulse center times.
    Returns np.array ref centers indexed by pulse-1.
    Same output format as original function.
    """

    # ---- TYPE YOUR REFERENCE CENTERS HERE ----
    center1 = 65.86585647
    center2 = 127.1104085
    center3 = 188.3631584
    center4 = 249.6256098
    # add more as needed

    ref = np.array([
        center1,
        center2,
        center3,
        center4
    ], dtype=float)

    return ref


def align_centers_to_reference(detected_centers, ref_centers, tol):
    """
    Nearest-neighbor, one-to-one assignment of detected centers to reference centers
    with max allowed time difference `tol`.
    Returns aligned array len(ref_centers) with np.nan where missing.
    """
    detected = np.asarray(detected_centers, dtype=float)
    ref = np.asarray(ref_centers, dtype=float)

    aligned = np.full(len(ref), np.nan, dtype=float)
    used_ref = set()

    valid_ref_idx = [i for i in range(len(ref)) if np.isfinite(ref[i])]
    if not valid_ref_idx:
        return aligned

    for c in detected:
        if not np.isfinite(c):
            continue

        best_i = None
        best_d = None
        for i in valid_ref_idx:
            if i in used_ref:
                continue
            d = abs(c - ref[i])
            if best_d is None or d < best_d:
                best_d = d
                best_i = i

        if best_i is not None and best_d is not None and best_d <= tol:
            aligned[best_i] = c
            used_ref.add(best_i)

    return aligned

def estimate_minMag_for_rx1_rx2(root_dir: Path,
                               n_folders: int = 10,
                               wn=(0.005, 0.3),
                               cutoff_time=50,
                               percentile=99.5,
                               reduction=0.35,
                                filtering = 0,
                                sigma = 3):
    """
    Auto-tune min magnitude separately for capture_1 (Rx1) and capture_2 (Rx2)
    by scanning the first N folders containing .iq files.
    Returns: (minMag_cap1, minMag_cap2)
    """
    # Find first N folders that contain any .iq
    folders = []

    #folders = [root_dir]

    for p in sorted(root_dir.rglob("*")):
        if p.is_dir() and any(p.glob("*.iq")):
            folders.append(p)
            if len(folders) >= n_folders:
                break


    if not folders:
        print("[AUTO] No .iq folders found. Using defaults minMag_cap1=minMag_cap2=2.0")
        return 2.0, 2.0

    def estimate_one_file(iq_path: Path, filtering = 0, sigma = 3) -> float | None:
        raw = np.fromfile(iq_path, dtype=np.int8)
        if raw.size < 4:
            return None

        I = raw[0::2].astype(np.float64)
        Q = raw[1::2].astype(np.float64)
        IQ = I + 1j * Q
        IQ -= np.mean(IQ)

        tt = np.arange(1, len(IQ) + 1) / 20.0  # keep your axis

        b, a = butter(N=4, Wn=wn, btype="bandpass")

        IQ = filtfilt(b, a, IQ)
        metadata = MetaDataObj()

        #IQ = filter_signal(metadata, IQ)
        if filtering == 0:
            mag = np.abs(IQ)
        elif filtering == 1:
            mag = gaussian_laplace(np.abs(IQ), sigma)
        elif filtering == 2:

            mag = gaussian_filter1d(np.abs(IQ), sigma)


        mag = mag[tt > cutoff_time]
        if mag.size == 0:
            return None

        return float(np.percentile(mag, percentile))

    cap1_ests = []
    cap2_ests = []

    for folder in folders:
        # Grab first file matching capture_1 and capture_2 in this folder (if present)
        cap1_files = sorted(folder.glob("*rx1*.iq"))
        cap2_files = sorted(folder.glob("*rx2*.iq"))

        if cap1_files:
            e1 = estimate_one_file(cap1_files[0],filtering, sigma)
            if e1 is not None:
                cap1_ests.append(e1)

        if cap2_files:
            e2 = estimate_one_file(cap2_files[0],filtering, sigma)
            if e2 is not None:
                cap2_ests.append(e2)

    # Fallback if missing one side
    def finalize(est_list, fallback):
        if not est_list:
            return fallback
        avg_est = float(np.mean(est_list))
        return avg_est * (1.0 - reduction)

    # If one side missing entirely, fall back to the other side (or 2.0)
    fallback_base = 2.0
    if cap1_ests:
        fallback_base = float(np.mean(cap1_ests)) * (1.0 - reduction)
    elif cap2_ests:
        fallback_base = float(np.mean(cap2_ests)) * (1.0 - reduction)

    minMag_cap1 = finalize(cap1_ests, fallback_base)
    minMag_cap2 = finalize(cap2_ests, fallback_base)

    print(f"[AUTO] Rx1/cap1: used {len(cap1_ests)} folders, minMag={minMag_cap1:.3f}")
    print(f"[AUTO] Rx2/cap2: used {len(cap2_ests)} folders, minMag={minMag_cap2:.3f}")

    return minMag_cap1, minMag_cap2



def compute_overall_avg_centers_with_gating(centers_rows, tol=10):
    """
    Align each run's detected pulse centers to a reference (per capture)
    so missing pulses don't shift indexing.
    Returns:
      matched[(cap, pulse_idx)] -> list of center times matched to that pulse
      ref1, ref2 reference arrays for capture 1 and 2
    """
    ref1 = build_reference_centers(centers_rows, capture_num=1)
    ref2 = build_reference_centers(centers_rows, capture_num=2)

    per_test = defaultdict(list)  # (folder, cap) -> list of (pulse_idx, center_time)
    for folder, cap, pulse, center_t, fname in centers_rows:
        per_test[(folder, int(cap))].append((int(pulse), float(center_t)))

    matched = defaultdict(list)

    for (folder, cap), items in per_test.items():
        items.sort(key=lambda x: x[0])
        detected_centers = [c for _, c in items]

        ref = ref1 if cap == 1 else ref2
        if ref.size == 0:
            continue

        aligned = align_centers_to_reference(detected_centers, ref, tol=tol)

        for i, c in enumerate(aligned, start=1):
            if np.isfinite(c):
                matched[(cap, i)].append(float(c))

    return matched, ref1, ref2

def parse_args_with_prompts():
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=str, default=None, help="Root directory to process")
    '''
    parser.add_argument("--minMag_cap1", type=float, default=None, help="Min peak height threshold for cap1")
    parser.add_argument("--minMag_cap2", type=float, default=None, help="Min peak height threshold for cap2")
    parser.add_argument("--cluster", type=int, default=None, help="Min points to qualify as a cluster")
    parser.add_argument("--clusterWeedOutDist", type=float, default=None, help="Max spacing within a cluster")
    parser.add_argument("--filtering", type=int, default=None, help="choice to filter")
    '''
    args = parser.parse_args()

    # Prompt if missing (keeps old behavior but interactive)
    '''
    if args.minMag_cap1 is None:
        val = input("Enter minimumMag1 (default 2): ").strip()
        args.minMag_cap1 = float(val) if val else 2.0

    if args.minMag_cap2 is None:
        val = input("Enter minimumMag2 (default 2): ").strip()
        args.minMag_cap2 = float(val) if val else 2.0

    if args.cluster is None:
        val = input("Enter cluster (default 7): ").strip()
        args.cluster = int(val) if val else 7

    if args.clusterWeedOutDist is None:
        val = input("Enter clusterWeedOutDist (default 3.5): ").strip()
        args.clusterWeedOutDist = float(val) if val else 3.5

    if args.filtering is None:
        val = input("Enter 1 for filtering or 0 for no filtering: ").strip()
        args.filtering = int(val) if val else 1
    '''
    root_dir = Path(args.root) if args.root else None


    return root_dir

def get_root_dir():
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=str, default=None, help="Root directory to process")
    args = parser.parse_args()

    if args.root:
        return Path(args.root)
    return None

def choose_folder_gui(title="Select a folder to process"):
    root = tk.Tk()
    root.withdraw()          # hide the main tkinter window
    root.attributes("-topmost", True)  # bring dialog to front

    folder = filedialog.askdirectory(title=title)
    root.destroy()

    return Path(folder) if folder else None

def process_one_iq_file(iq_path: Path,
                        fs=20e6,
                        wn=(0.005, 0.3),
                        minimumMag=2,
                        cutoff_time=50,
                        cluster=7,
                        clusterWeedOutDist=3.5,
                        timeFilter=350,
                        timeBetweenClusters=30,
                        gap_for_edges=5,
                        filtering = 0,
                        sigma = 3):
    """
    Processes a single .iq file and returns:
      - tt, IQ_data
      - clusterLocsArray, clusterPeaksArray (red points)
      - ToFtimesArray, ToFpeaksArray (blue edge points)
      - CalcToFArray (ToF per pulse)
    """
    '''
    # --- IMPORTANT: open IQ file as BINARY ---
    raw_data = np.fromfile(iq_path, dtype=np.int8)
    if raw_data.size < 4:
        return None


    # Interleaved IQ
    I = raw_data[0::2]
    Q = raw_data[1::2]
    IQ_data = I.astype(np.float64) + 1j * Q.astype(np.float64)

    # DC offset removal
    IQ_data = IQ_data - np.mean(IQ_data)
    '''


    raw_data = np.fromfile(iq_path, dtype=np.int8)
    IQ_data = raw_data.astype(np.float32).view(np.complex64)
    IQ_data = IQ_data - np.mean(IQ_data)

    bandstop_taps = firwin(numtaps=301, cutoff=[0.3, 0.99], pass_zero=True)
    IQ_data = lfilter(bandstop_taps, 1.0, IQ_data)

    # Keep your original time axis definition (matches your prior code)
    tt = np.arange(1, len(IQ_data) + 1) / 20.0
    metadata = MetaDataObj()
    # Butter bandpass + filtfilt (MATLAB-style)
    b, a = butter(N=4, Wn=wn, btype='bandpass')
    IQ_data = filtfilt(b, a, IQ_data)


    #IQ_data = filter_signal(metadata, IQ_data)


    # Peak detection on magnitude
    if filtering == 0:
        mag = np.abs(IQ_data)
    elif filtering == 1:
        mag = gaussian_laplace(np.abs(IQ_data), sigma)
    elif filtering == 2:

        mag = gaussian_filter1d(np.abs(IQ_data), sigma)

    idx, props = find_peaks(mag, height=minimumMag)
    peaks = mag[idx]
    locs = tt[idx]

    # Ignore startup peaks
    valid = locs > cutoff_time
    locs = locs[valid]
    peaks = peaks[valid]

    # ---------- CLUSTERING (no duplicates) ----------
    clusterLocsArray = []
    clusterPeaksArray = []

    s = 0
    endDat = len(locs) - 1
    while s < endDat:
        if (locs[s + 1] - locs[s]) < clusterWeedOutDist:
            startDel = s

            clusterCount = 0
            c = s
            while c < endDat and (locs[c + 1] - locs[c]) < clusterWeedOutDist:
                clusterCount += 1
                c += 1

            if clusterCount >= cluster:
                endDel = startDel + clusterCount

                for k in range(startDel, endDel + 1):
                    if k >= len(locs):
                        break
                    if locs[k] > timeFilter:
                        break

                    clusterLocsArray.append(locs[k])
                    clusterPeaksArray.append(peaks[k])

                s = endDel + 1
                continue

        s += 1

    clusterLocsArray = np.array(clusterLocsArray)
    clusterPeaksArray = np.array(clusterPeaksArray)

    pulseCentersArray = pulse_centers_from_clustered_peaks(
        clusterLocsArray,
        clusterPeaksArray,
        timeBetweenClusters=timeBetweenClusters,
        weighted=True
    )

    # ---------- EDGE PICKING (your blue points logic, made robust) ----------
    # Instead of fixed-size arrays (np.zeros(8)), build dynamically:
    ToFtimes = []
    ToFpeaks = []

    if len(clusterLocsArray) >= 2:
        first_idx = 0
        for i in range(len(clusterLocsArray) - 1):
            # gap indicates next pulse cluster
            if (clusterLocsArray[i + 1] - clusterLocsArray[i]) > gap_for_edges:
                last_idx = i

                # store edge points: first and last of this cluster group
                ToFtimes.extend([clusterLocsArray[first_idx], clusterLocsArray[last_idx]])
                ToFpeaks.extend([clusterPeaksArray[first_idx], clusterPeaksArray[last_idx]])

                first_idx = i + 1

        # finalize last group
        last_idx = len(clusterLocsArray) - 1
        ToFtimes.extend([clusterLocsArray[first_idx], clusterLocsArray[last_idx]])
        ToFpeaks.extend([clusterPeaksArray[first_idx], clusterPeaksArray[last_idx]])

    ToFtimesArray = np.array(ToFtimes)
    ToFpeaksArray = np.array(ToFpeaks)

    # ---------- ToF calculations (pairwise differences) ----------
    CalcToF = []
    for i in range(0, len(ToFtimesArray) - 1, 2):
        CalcToF.append(ToFtimesArray[i + 1] - ToFtimesArray[i])
    CalcToFArray = np.array(CalcToF)

    return {
        "tt": tt,
        "IQ_data": IQ_data,
        "clusterLocsArray": clusterLocsArray,
        "clusterPeaksArray": clusterPeaksArray,
        "ToFtimesArray": ToFtimesArray,
        "ToFpeaksArray": ToFpeaksArray,
        "CalcToFArray": CalcToFArray,
        "pulseCentersArray": pulseCentersArray,
        "picked_idx": clusterLocsArray,
        "picked_vals": clusterPeaksArray,
    }

def pulse_centers_from_clustered_peaks(clusterLocsArray,
                                       clusterPeaksArray,
                                       timeBetweenClusters=50,
                                       weighted=True):
    """
    Groups clustered peaks into pulses using timeBetweenClusters gap.
    Returns one center time per pulse.
    """
    locs = np.asarray(clusterLocsArray)
    peaks = np.asarray(clusterPeaksArray)

    if locs.size == 0:
        return np.array([])

    centers = []
    start = 0

    for i in range(locs.size - 1):
        if (locs[i + 1] - locs[i]) > timeBetweenClusters:
            end = i
            seg_locs = locs[start:end + 1]
            seg_peaks = peaks[start:end + 1]

            if weighted and np.sum(seg_peaks) > 0:
                center = float(np.sum(seg_locs * seg_peaks) / np.sum(seg_peaks))
            else:
                center = float(np.mean(seg_locs))

            centers.append(center)
            start = i + 1

    # finalize last group
    end = locs.size - 1
    seg_locs = locs[start:end + 1]
    seg_peaks = peaks[start:end + 1]

    if weighted and np.sum(seg_peaks) > 0:
        center = float(np.sum(seg_locs * seg_peaks) / np.sum(seg_peaks))
    else:
        center = float(np.mean(seg_locs))

    centers.append(center)
    return np.array(centers, dtype=float)

def compute_overall_avg_centers_with_baseline(centers_rows, baseline_ref, tol=2):
    """
    Align every run's detected pulse centers to a fixed baseline reference.
    Prevents pulse index shifting when a pulse is missed.
    """
    per_test = defaultdict(list)  # (folder, cap) -> list of (pulse_idx, center_time)
    for folder, cap, pulse, center_t, fname in centers_rows:
        per_test[(folder, int(cap))].append((int(pulse), float(center_t)))

    matched = defaultdict(list)   # (cap, pulse_idx) -> list of matched center times

    for (folder, cap), items in per_test.items():
        items.sort(key=lambda x: x[0])
        detected_centers = [c for _, c in items]

        aligned = align_centers_to_reference(detected_centers, baseline_ref, tol=tol)

        for i, c in enumerate(aligned, start=1):
            if np.isfinite(c):
                matched[(cap, i)].append(float(c))

    return matched


def write_pulse_centers_excel(xlsx_path: Path, centers_rows):
    """
    centers_rows: list of tuples (Folder, Capture, Pulse, CenterTime, FileName)

    Creates a workbook with:
      Sheet1: Per Test Pulse Centers (WIDE)  -> one row per Folder, columns per pulse grouped by capture
      Sheet2: Overall Avg/Median Pulse Centers (by Pulse #)  -> keep your existing summary table style
    """
    wb = Workbook()

    # Baseline centers (your wired/reference capture baseline)
    BASELINE_C2 = np.array([
        65.86585647,
        127.1104085,
    188.3631584,
    249.6256098
    ], dtype=float)


    K = len(BASELINE_C2)
    TOLERANCE = 7

    # ---------------- Sheet 1: per test (WIDE) ----------------
    ws1 = wb.active
    ws1.title = "Per Test Pulse Centers"

    headers1 = (["Folder"]
                + [f"C1_P{p}" for p in range(1, K + 1)]
                + [f"C2_P{p}" for p in range(1, K + 1)])
    ws1.append(headers1)
    for c in range(1, len(headers1) + 1):
        ws1.cell(row=1, column=c).font = Font(bold=True)

    # Build per-folder detected centers per capture
    per_test = defaultdict(lambda: {1: [], 2: []})  # folder -> {cap -> list of detected centers}

    # Keep pulse-ordering inside each file/run; we only need centers (not pulse indices) for alignment
    # because align_centers_to_reference does 1-to-1 NN assignment vs BASELINE_C2.
    temp = defaultdict(list)  # (folder, cap) -> list of (pulse_idx, center_time)
    for folder, cap, pulse, center_t, fname in centers_rows:
        cap = int(cap)
        if cap not in (1, 2):
            continue
        temp[(folder, cap)].append((int(pulse), float(center_t)))

    # Align each folder+capture to baseline so columns stay stable even if a pulse is missed
    aligned_by_folder = defaultdict(lambda: {1: np.full(K, np.nan), 2: np.full(K, np.nan)})

    for (folder, cap), items in temp.items():
        items.sort(key=lambda x: x[0])              # sort by pulse index
        detected_centers = [c for _, c in items]    # just center times in pulse order
        aligned = align_centers_to_reference(detected_centers, BASELINE_C2, tol=TOLERANCE)  # len K
        aligned_by_folder[folder][cap] = aligned

    # Write rows (one row per folder)
    for folder in sorted(aligned_by_folder.keys()):
        row = [folder]

        for cap in (1, 2):
            arr = aligned_by_folder[folder][cap]
            for v in arr:
                row.append("" if not np.isfinite(v) else float(v))

        ws1.append(row)

    ws1.freeze_panes = "A2"

    # autosize columns (basic)
    for col in range(1, len(headers1) + 1):
        max_len = 10
        for r in range(1, ws1.max_row + 1):
            v = ws1.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws1.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    # ---------------- Sheet 2: overall summary (keep your current style) ----------------
    ws2 = wb.create_sheet("Overall Median Pulse Centers")

    headers2 = ["Capture", "Pulse", "MedianCenterTime", "N", "ReferenceCenter", "ToleranceUsed"]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)

    matched = compute_overall_avg_centers_with_baseline(centers_rows, BASELINE_C2, tol=TOLERANCE)

    for cap in (1, 2):
        for pulse_idx in range(1, K + 1):
            vals = matched.get((cap, pulse_idx), [])
            # use median if that's what you want; swap to np.mean(vals) if you prefer average
            avg = float(np.median(vals)) if vals else ""
            n = len(vals)
            ws2.append([cap, pulse_idx, avg, n, float(BASELINE_C2[pulse_idx - 1]), TOLERANCE])

    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 14

    wb.save(xlsx_path)


def save_plot(out, iq_path: Path, filtering = 0, sigma = 3):
    """
    Saves a plot of |IQ| vs sample index with clustered (final) peaks overlaid.

    Uses:
      - out["IQ_data"]
      - out["clusterLocsArray"]  (peak x positions; either sample index OR tt-units)
      - out["clusterPeaksArray"] (peak magnitudes)
    """
    IQ_data = out["IQ_data"]


    clusterLocsArray = out.get("clusterLocsArray", np.array([]))
    clusterPeaksArray = out.get("clusterPeaksArray", np.array([]))

    if filtering == 0:
        mag = np.abs(IQ_data)
    elif filtering == 1:
        mag = gaussian_laplace(np.abs(IQ_data), sigma)
    elif filtering == 2:

        mag = gaussian_filter1d(np.abs(IQ_data), sigma)
    x = out["tt"]

    # ---- Convert cluster x locations to sample indices if needed ----
    # If clusterLocsArray looks like time (floats like 50.2, 120.7, etc),
    # convert using your tt scaling (tt = (idx+1)/20 in your original code).
    cluster_x = np.array(clusterLocsArray)

    if clusterLocsArray.size > 0:
        plt.scatter(clusterLocsArray, clusterPeaksArray, c="red", s=28, label="Picked Peaks", zorder=5)

    # ---- Plot ----
    plt.figure(figsize=(11, 4))
    plt.plot(x, mag, label="|IQ|", alpha=0.8)

    if len(cluster_x) > 0:
        plt.scatter(cluster_x, clusterPeaksArray, c="red", s=28, label="Picked Peaks", zorder=5)

    plt.title(f"{iq_path.parent.name} / {iq_path.name}")
    plt.xlabel("MicroSeconds")
    plt.ylabel("Magnitude")
    plt.grid(True, which="both", linestyle="--", alpha=0.35)
    plt.minorticks_on()
    plt.legend()
    plt.tight_layout()

    out_png = iq_path.parent / f"{iq_path.stem}_peaks.png"
    plt.savefig(out_png, dpi=200)
    plt.close()
    return out_png


def save_tof_txt(out, iq_path: Path):
    CalcToFArray = out["CalcToFArray"]

    out_txt = iq_path.parent / f"{iq_path.stem}_tof.txt"
    with open(out_txt, "w", encoding="utf-8") as f:
        f.write(f"File: {iq_path}\n")
        f.write(f"Num ToF pulses: {len(CalcToFArray)}\n\n")
        for i, tof in enumerate(CalcToFArray, start=1):
            f.write(f"The time of flight for signal {i} is {tof}\n")
    return out_txt

def load_existing_rows(csv_path: Path):
    rows = {}  # key: (Folder, Capture, Pulse) -> ToF
    if csv_path.exists():
        with open(csv_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                key = (r["Folder"], int(r["Capture"]), int(r["Pulse"]))
                rows[key] = float(r["ToF"])
    return rows

def upsert_tofs(rows_dict, run_name, capture_num, tof_array):
    for pulse_idx, tof in enumerate(tof_array, start=1):
        key = (run_name, int(capture_num), int(pulse_idx))
        rows_dict[key] = float(tof)  # replace if exists, add if not

def write_rows_atomic(csv_path: Path, rows_dict):
    tmp_path = csv_path.with_suffix(".tmp")
    with open(tmp_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Folder", "Capture", "Pulse", "ToF"])

        # stable output ordering
        for (folder, cap, pulse) in sorted(rows_dict.keys()):
            writer.writerow([folder, cap, pulse, rows_dict[(folder, cap, pulse)]])

    tmp_path.replace(csv_path)  # atomic-ish replace on Windows

def infer_run_name(iq_path: Path) -> str:
    # Find the nearest parent folder that looks like run_#### (robust)
    for p in iq_path.parents:
        if p.name.lower().startswith("run_"):
            return p.name
    return iq_path.parent.name  # fallback

def infer_capture_num(iq_path: Path) -> int:
    # Parses "..._capture_1.iq" safely (avoids grabbing things like 0897)
    tokens = (iq_path.stem)
    #.split("_"))
    print(tokens[2])
    '''
    for i, tok in enumerate(tokens[:-1]):
        if tok.lower() == "rx" and tokens[i + 1].isdigit():
            return int(tokens[i + 1])
    '''
    return int(tokens[2])  # fallback

def interactive_pick_points(x, y, title="Pick points", snap=True):
    """
    Interactive picker:
      - Arrow keys move a cursor along x/y (sample-by-sample)
      - Enter/Space picks the current cursor point
      - Left click picks nearest point to mouse
      - 'u' undo last pick
      - 's' save picks (returns list)
      - 'q' or Esc closes

    Returns: list of (x_pick, y_pick)
    """
    x = np.asarray(x)
    y = np.asarray(y)

    picked = []
    idx = 0

    fig, ax = plt.subplots(figsize=(11, 4))
    ax.plot(x, y, alpha=0.85, label="signal")
    ax.set_title(title)
    ax.grid(True, which="both", linestyle="--", alpha=0.35)
    ax.minorticks_on()

    # Cursor visuals
    vline = ax.axvline(x[idx], linewidth=1.0)
    cursor_pt, = ax.plot([x[idx]], [y[idx]], marker="o", markersize=6, linestyle="None")

    # Picked visuals
    picked_scatter = ax.scatter([], [], s=40, label="picked", zorder=5)
    info = ax.text(
        0.01, 0.98, "", transform=ax.transAxes, va="top", ha="left"
    )

    def _update_cursor():
        nonlocal idx
        idx = int(np.clip(idx, 0, len(x) - 1))
        vline.set_xdata([x[idx], x[idx]])
        cursor_pt.set_data([x[idx]], [y[idx]])
        info.set_text(
            f"idx={idx}  x={x[idx]:.6f}  y={y[idx]:.6f}\n"
            f"picked={len(picked)}  (←/→ move, Enter pick, click pick, u undo, s save, q quit)"
        )
        fig.canvas.draw_idle()

    def _refresh_picks():
        if picked:
            px = [p[0] for p in picked]
            py = [p[1] for p in picked]
        else:
            px, py = [], []
        picked_scatter.set_offsets(np.c_[px, py] if len(px) else np.empty((0, 2)))
        fig.canvas.draw_idle()

    def _pick_at_index(i):
        picked.append((float(x[i]), float(y[i])))
        _refresh_picks()

    def _nearest_index(xq):
        # fast nearest by x (assumes x is monotonic; yours is tt so it is)
        i = int(np.searchsorted(x, xq))
        if i <= 0:
            return 0
        if i >= len(x):
            return len(x) - 1
        # choose closer of i and i-1
        return i if abs(x[i] - xq) < abs(x[i - 1] - xq) else i - 1

    def on_key(event):
        nonlocal idx
        if event.key in ("right", "d"):
            idx += 1
            _update_cursor()
        elif event.key in ("left", "a"):
            idx -= 1
            _update_cursor()
        elif event.key in ("up", "w"):
            idx += 50
            _update_cursor()
        elif event.key in ("down", "s"):
            idx -= 50
            _update_cursor()
        elif event.key in ("enter", " "):
            _pick_at_index(idx)
        elif event.key == "u":
            if picked:
                picked.pop()
                _refresh_picks()
        elif event.key in ("escape", "q"):
            plt.close(fig)

    def on_click(event):
        nonlocal idx
        if event.inaxes != ax:
            return
        if event.button != 1:  # left-click only
            return
        # pick nearest by x
        i = _nearest_index(event.xdata)
        idx = i
        _update_cursor()
        _pick_at_index(i)

    fig.canvas.mpl_connect("key_press_event", on_key)
    fig.canvas.mpl_connect("button_press_event", on_click)

    ax.legend(loc="upper right")
    _update_cursor()
    plt.tight_layout()
    #plt.show()

    return picked


def save_manual_picks_csv(iq_path: Path, picked, out_name_suffix="_manual_picks.csv"):
    """
    Save picked points to a CSV next to the iq file.
    """
    out_csv = iq_path.parent / f"{iq_path.stem}{out_name_suffix}"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["x", "y"])
        for x, y in picked:
            w.writerow([x, y])
    return out_csv

def bulk_run(root_dir: Path, cluster, clusterWeedOutDist, filtering, sigma):
    if root_dir is None:
        print("No folder selected. Exiting.")
        return

    ROOT_DIR = Path(root_dir)
    centers_rows = []  # (Folder, Capture, Pulse, CenterTime, FileName)

    CSV_PATH = ROOT_DIR / "tof_results.csv"
    rows = load_existing_rows(CSV_PATH)  # existing upsert map (Folder,Capture,Pulse)->ToF

    minMag_cap1, minMag_cap2 = estimate_minMag_for_rx1_rx2(ROOT_DIR, n_folders=10,filtering=filtering, sigma = sigma)

    iq_files = sorted(ROOT_DIR.rglob("*.iq"))
    print(f"Selected folder: {ROOT_DIR}")
    print(f"Found {len(iq_files)} .iq files")

    for iq_path in iq_files:
        run_name = infer_run_name(iq_path)
        capture_num = infer_capture_num(iq_path)

        print("this is the capture num: ", capture_num, "from run: ", run_name)

        if capture_num == 1:
            minimumMag = 1
        elif capture_num == 2:
            minimumMag = 33
        else:
            minimumMag = minMag_cap1  # fallback (or skip)



        out = process_one_iq_file(
            iq_path,
            minimumMag=minimumMag,
            cluster=cluster,
            clusterWeedOutDist=clusterWeedOutDist,
            filtering=filtering,
            sigma = sigma,

        )
        if out is None:
            print(f"[SKIP] {iq_path} (empty/too small)")
            continue

        centers = out.get("pulseCentersArray", np.array([]))
        for pulse_idx, center_t in enumerate(centers, start=1):
            centers_rows.append((run_name, int(capture_num), int(pulse_idx), float(center_t), iq_path.name))

        # Optional guard: only allow capture 1/2 (prevents weird 4-digit capture IDs)
        # If you sometimes have capture_3 etc, expand this tuple.
        if capture_num not in (1, 2):
            print(f"[WARN] capture_num={capture_num} (unexpected) for {iq_path.name} -> skipping CSV write")
        else:
            # ✅ ONE source of truth for CSV writing (upsert then rewrite)
            upsert_tofs(rows, run_name, capture_num, out["CalcToFArray"])
            write_rows_atomic(CSV_PATH, rows)

        # Console output
        for i, tof in enumerate(out["CalcToFArray"], start=1):
            print(f"{iq_path.name}: The time of flight for signal {i} is {tof}")

        # Save plot + txt INSIDE the same folder as the iq file (once)


        # ---- interactive/manual picking ----
        # Recompute the plotted magnitude exactly like save_plot does
        IQ_data = out["IQ_data"]
        if filtering == 0:
            mag = np.abs(IQ_data)
        elif filtering == 1:
            mag = gaussian_laplace(np.abs(IQ_data), sigma)
        elif filtering == 2:
            mag = gaussian_filter1d(np.abs(IQ_data), sigma)

        x = out["tt"]

        # Show interactive picker
        picked = interactive_pick_points(
            x, mag,
            title=f"{iq_path.parent.name} / {iq_path.name}\n(click to pick, arrows to move, Enter pick, u undo, q quit)"
        )

        # Save picks (if any)
        if picked:
            out_csv = save_manual_picks_csv(iq_path, picked)
            print(f"[MANUAL] Saved {len(picked)} picks -> {out_csv.name}")
        else:
            print("[MANUAL] No picks made.")

        # Keep your existing outputs too
        png_path = save_plot(out, iq_path, filtering, sigma)
        txt_path = save_tof_txt(out, iq_path)
        print(f"[OK] {iq_path.name} -> {txt_path.name}, {png_path.name}")

        html_path = save_interactive_html_clickpick(x, mag, iq_path)
        print(f"[INTERACTIVE SAVED] {html_path}")

    xlsx_path = ROOT_DIR / "pulse_center_times.xlsx"
    write_pulse_centers_excel(xlsx_path, centers_rows)
    print(f"[OK] Wrote pulse centers Excel: {xlsx_path}")

    print(f"\nDone. CSV: {CSV_PATH}")

if __name__ == "__main__":
    root_dir = parse_args_with_prompts()
    #Path("/opt/TVWS/Data/Latest/"))


    cluster = 7
    clusterWeedOutDist= 3.5
    filtering = 0
    sigma = 4


    if root_dir is None:
        root_dir = choose_folder_gui("Pick the folder to iterate over (.iq bulk folder)")
        if root_dir is None:
            print("No folder selected. Exiting.")
            raise SystemExit(1)

    bulk_run(root_dir, cluster, clusterWeedOutDist, filtering, sigma)
